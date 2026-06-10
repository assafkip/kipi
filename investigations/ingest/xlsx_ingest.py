"""Excel (.xlsx) ingestion. Flattens rows to text like csv_ingest."""
from pathlib import Path


def extract_text(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    out: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out.append(f"\n=== SHEET: {sheet_name} ===")
        headers: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c).strip() if c is not None else "" for c in row]
                continue
            cells = []
            for j, cell in enumerate(row):
                if cell is None or cell == "":
                    continue
                header = headers[j] if j < len(headers) else f"col{j}"
                cells.append(f"{header}: {cell}")
            if cells:
                out.append(" | ".join(cells))
    wb.close()
    return "\n".join(out)
