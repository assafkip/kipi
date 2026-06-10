"""Structured CSV/TSV ingest — turn columns into TYPED entities, deterministically.

A spreadsheet is not prose. A column of wallet addresses, a column of emails, a
column of domains — the regex-over-flat-text path misses most of these (it found
0 wallets in a wallet-fraud case). Here each COLUMN is typed once (by header
name + by running its values through the same extractor regexes), then every
cell in a typed column becomes a typed entity with the row as its context.

Design choices (the panel flagged the risks):
- ONE `dataset` report per file (evidence_kind='dataset'), not one report per row.
  Avoids report explosion + delete_report loops.
- The dataset report's raw_text is a BOUNDED summary (header + inferred column
  types + a row preview), so a 50k-row CSV can't blow the 60k Understand budget.
- Rows are capped (MAX_ROWS) and the cap is LOGGED into the report (no silent
  truncation).
- Fully deterministic. No LLM. Columns the regexes can't type stay as context.
"""
import csv as csvmod
import io
from pathlib import Path

from investigations.storage import db
from investigations.ingest import extractor as ext

MAX_ROWS = 5000
SAMPLE_ROWS = 50          # rows sampled to infer each column's type
PREVIEW_ROWS = 20         # rows shown in the dataset report's raw_text

# Column-name hints (substring → surface type). Used when the values don't
# self-identify (a 'name' column has no regex signature).
HEADER_HINTS = [
    ("wallet", "crypto_wallet"), ("address", "crypto_wallet"),
    ("email", "email"), ("e-mail", "email"),
    ("domain", "domain"), ("website", "domain"), ("site", "domain"),
    ("url", "url"), ("link", "url"),
    ("ipv4", "ip"), ("ip_", "ip"), ("ip address", "ip"),
    ("phone", "phone"), ("mobile", "phone"), ("tel", "phone"),
    ("username", "handle"), ("handle", "handle"), ("screen_name", "handle"),
    ("sha256", "hash_sha256"), ("sha", "hash_sha256"), ("md5", "hash_md5"),
    ("full name", "person"), ("name", "person"),
]

# Value matchers (surface type → predicate). Order = priority on ambiguous cells.
VALUE_MATCHERS = [
    ("crypto_wallet", lambda v: bool(ext.WALLET_RE.fullmatch(v))),
    ("email", lambda v: bool(ext.EMAIL_RE.fullmatch(v))),
    ("ip", lambda v: bool(ext.IPV4_RE.fullmatch(v))),
    ("hash_sha256", lambda v: bool(ext.SHA256_RE.fullmatch(v))),
    ("hash_md5", lambda v: bool(ext.MD5_RE.fullmatch(v))),
    ("url", lambda v: bool(ext.URL_RE.fullmatch(v))),
    ("domain", lambda v: bool(ext.DOMAIN_RE.fullmatch(v))),
]


def _read_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    if path.suffix.lower() in (".xlsx", ".xls"):
        return _read_rows_xlsx(path)
    text = path.read_text(encoding="utf-8", errors="replace")
    delim = "\t" if (path.suffix.lower() == ".tsv" or text[:2000].count("\t") > text[:2000].count(",")) else ","
    reader = csvmod.reader(io.StringIO(text), delimiter=delim)
    rows = [r for r in reader if any(c.strip() for c in r)]
    if not rows:
        return [], []
    header = [h.strip() for h in rows[0]]
    return header, rows[1:]


def _read_rows_xlsx(path: Path) -> tuple[list[str], list[list[str]]]:
    """Header + rows from the FIRST sheet that has data. Cells coerced to str."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            grid = []
            for row in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c).strip() for c in row]
                if any(cells):
                    grid.append(cells)
            if len(grid) >= 2:                       # a header + at least one data row
                width = max(len(r) for r in grid)
                grid = [r + [""] * (width - len(r)) for r in grid]
                return [h.strip() for h in grid[0]], grid[1:]
        return [], []
    finally:
        wb.close()


def _header_type(name: str) -> str | None:
    low = name.lower()
    for frag, t in HEADER_HINTS:
        if frag in low:
            return t
    return None


def _value_type(value: str) -> str | None:
    v = value.strip()
    if not v:
        return None
    for t, pred in VALUE_MATCHERS:
        try:
            if pred(v):
                return t
        except Exception:
            continue
    return None


def _profile_columns(header: list[str], rows: list[list[str]]) -> dict[int, str]:
    """col index → surface type. Strong value-majority wins; else header hint."""
    col_types: dict[int, str] = {}
    for ci, hname in enumerate(header):
        sampled = [r[ci] for r in rows[:SAMPLE_ROWS] if ci < len(r) and r[ci].strip()]
        vt_counts: dict[str, int] = {}
        for val in sampled:
            t = _value_type(val)
            if t:
                vt_counts[t] = vt_counts.get(t, 0) + 1
        if sampled and vt_counts:
            best_t = max(vt_counts, key=vt_counts.get)
            if vt_counts[best_t] / len(sampled) >= 0.5:
                col_types[ci] = best_t
                continue
        ht = _header_type(hname)
        if ht:
            col_types[ci] = ht
    return col_types


def _row_context(header: list[str], row: list[str]) -> str:
    parts = []
    for ci, h in enumerate(header):
        if ci < len(row) and row[ci].strip():
            parts.append(f"{h}={row[ci].strip()}")
    return " | ".join(parts)[:300]


def _summary(path: Path, header: list[str], rows: list[list[str]],
             col_types: dict[int, str], capped: int) -> str:
    lines = [f"# Dataset: {path.name}", "",
             f"{len(rows)} data rows, {len(header)} columns.",
             ("Typed up to %d rows." % capped) if capped else "All rows typed.",
             "", "## Columns"]
    for ci, h in enumerate(header):
        t = col_types.get(ci)
        lines.append(f"- {h}: {t if t else 'context'}")
    lines += ["", "## Preview"]
    lines.append(" | ".join(header))
    for r in rows[:PREVIEW_ROWS]:
        lines.append(" | ".join((c or "").strip() for c in r))
    return "\n".join(lines)


def ingest(conn, path: Path, file_hash: str, investigation: str | None) -> dict | None:
    """Structured ingest of a delimited file. Returns {report_id, ...} or None if
    the file has no usable rows."""
    header, rows = _read_rows(path)
    if not header or not rows:
        return None
    col_types = _profile_columns(header, rows)
    capped = MAX_ROWS if len(rows) > MAX_ROWS else 0
    use_rows = rows[:MAX_ROWS]

    summary = _summary(path, header, rows, col_types, capped)
    source_type = "xlsx" if path.suffix.lower() in (".xlsx", ".xls") else "csv"
    report_id = db.insert_report(conn, str(path), file_hash, source_type,
                                 path.stem, investigation, summary)
    conn.execute("UPDATE reports SET evidence_kind = 'dataset' WHERE id = ?", (report_id,))

    entities_added, mentions = 0, 0
    seen_in_report: set[tuple[int, str]] = set()
    for row in use_rows:
        ctx = _row_context(header, row)
        for ci, surface in col_types.items():
            if ci >= len(row):
                continue
            value = row[ci].strip()
            if not value:
                continue
            # person columns: keep only plausible multi-word names, skip ids/blanks.
            if surface == "person" and not _looks_like_name(value):
                continue
            key = (ci, value.lower())
            if key in seen_in_report:
                continue
            seen_in_report.add(key)
            eid = db.upsert_entity(conn, value, surface, report_id)
            conn.execute("UPDATE entities SET case_type = COALESCE(case_type, ?) WHERE id = ?",
                         (surface, eid))
            db.add_mention(conn, eid, report_id, value, ctx)
            mentions += 1
            entities_added += 1
    conn.commit()
    typed_cols = sum(1 for _ in col_types)
    return {"report_id": report_id, "rows": len(use_rows), "row_total": len(rows),
            "typed_columns": typed_cols, "entities": entities_added,
            "mentions": mentions, "capped": bool(capped)}


def _looks_like_name(value: str) -> bool:
    parts = value.split()
    if not (1 < len(parts) <= 4):
        return False
    return all(p[:1].isalpha() for p in parts) and not any(ch.isdigit() for ch in value)
